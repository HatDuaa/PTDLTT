import openpyxl, csv, io
src=r"C:\Users\loocn\Documents\github\PTDLTT\60.xlsx"
wb = openpyxl.load_workbook(src, read_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    with open(f"{name}.csv","w",newline="",encoding="utf-8") as f:
        w=csv.writer(f)
        n=0
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if c is None else c for c in row]); n+=1
    print(name, n)
